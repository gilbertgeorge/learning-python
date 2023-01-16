from openpyxl import Workbook, load_workbook


def create_workbook():
    book = Workbook()
    new_sheet = book.active
    new_sheet['A1'] = 'cat'
    new_sheet['A2'] = 'dog'
    new_sheet['A3'] = 'tiger'
    new_sheet['A4'] = 'hamster'
    new_sheet.cell(row=5, column=1).value = 'iguana'

    book.save(r"..\supplemental\excel\New_file.xlsx")


def read_workbook():
    book = load_workbook(r"..\supplemental\excel\New_file.xlsx")
    # all_sheets_names = book.sheetnames
    # sheet = book[all_sheets_names[0]]
    sheet = book.active
    # print(sheet['A1'].value)

    rows = sheet.max_row
    cols = sheet.max_column
    # for row in range(1, rows + 1):
    #     string = ''
    #     for column in range(1, cols + 1):
    #         value = sheet.cell(row=row, column=column).value
    #         string = string + str(value) + ', '
    #     string = string[:-2]  # we remove the last appended comma and space
    #     print(string)
    for row in range(1, rows + 1):
        string = ''
        for column in range(1, cols + 1):
            value = sheet.cell(row=row, column=column).value
            string = string + str(value) + ', '
        string = string[:-2]
        print(string)


if __name__ == '__main__':
    # create_workbook()
    read_workbook()
